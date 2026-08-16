import sqlite3
import hashlib
import os
import random

import resend

from openpyxl import Workbook, load_workbook


DB_PATH = "users.db"
EXCEL_PATH = "users_data.xlsx"


# ==========================================================
# DATABASE
# ==========================================================

def get_connection():
    return sqlite3.connect(DB_PATH)


def init_db():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS login_activity (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            email TEXT NOT NULL,
            login_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.commit()
    conn.close()

    sync_excel()


# ==========================================================
# PASSWORD
# ==========================================================

def hash_password(password):

    return hashlib.sha256(
        password.encode("utf-8")
    ).hexdigest()


# ==========================================================
# EXCEL
# ==========================================================

def sync_excel():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, email, created_at
        FROM users
        ORDER BY created_at DESC
    """)

    users = cursor.fetchall()

    cursor.execute("""
        SELECT id, user_id, email, login_time
        FROM login_activity
        ORDER BY login_time DESC
    """)

    logins = cursor.fetchall()

    conn.close()

    if os.path.exists(EXCEL_PATH):

        wb = load_workbook(EXCEL_PATH)

    else:

        wb = Workbook()

        default_sheet = wb.active
        wb.remove(default_sheet)

    # ------------------------------------------------------
    # USERS SHEET
    # ------------------------------------------------------

    if "Users" in wb.sheetnames:

        ws = wb["Users"]
        ws.delete_rows(1, ws.max_row)

    else:

        ws = wb.create_sheet("Users")

    ws.append([
        "User ID",
        "Email",
        "Created At",
    ])

    for user in users:
        ws.append(user)

    # ------------------------------------------------------
    # LOGIN ACTIVITY SHEET
    # ------------------------------------------------------

    if "Login Activity" in wb.sheetnames:

        ws = wb["Login Activity"]
        ws.delete_rows(1, ws.max_row)

    else:

        ws = wb.create_sheet("Login Activity")

    ws.append([
        "Activity ID",
        "User ID",
        "Email",
        "Login Time",
    ])

    for login in logins:
        ws.append(login)

    # ------------------------------------------------------
    # DELETED USERS SHEET
    # ------------------------------------------------------

    if "Deleted Users" not in wb.sheetnames:

        ws = wb.create_sheet("Deleted Users")

        ws.append([
            "User ID",
            "Email",
            "Deleted At",
        ])

    # ------------------------------------------------------
    # COLUMN WIDTHS
    # ------------------------------------------------------

    for sheet in wb.worksheets:

        for column in sheet.columns:

            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                50
            )

    wb.save(EXCEL_PATH)


# ==========================================================
# REGISTER
# ==========================================================

def register_user(email, password):

    email = email.strip().lower()

    if not email or not password:

        return False, "Email and password are required."

    password_hash = hash_password(password)

    try:

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute(
            """
            INSERT INTO users (
                email,
                password_hash
            )
            VALUES (?, ?)
            """,
            (
                email,
                password_hash,
            )
        )

        conn.commit()
        conn.close()

        sync_excel()

        return True, "Account created successfully."

    except sqlite3.IntegrityError:

        return False, (
            "An account with this email already exists."
        )


# ==========================================================
# LOGIN
# ==========================================================

def login_user(email, password):

    email = email.strip().lower()

    password_hash = hash_password(password)

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute(
        """
        SELECT id, email
        FROM users
        WHERE email = ?
        AND password_hash = ?
        """,
        (
            email,
            password_hash,
        )
    )

    user = cursor.fetchone()

    if not user:

        conn.close()

        return False, "Invalid email or password."

    user_id = user[0]
    user_email = user[1]

    cursor.execute(
        """
        INSERT INTO login_activity (
            user_id,
            email
        )
        VALUES (?, ?)
        """,
        (
            user_id,
            user_email,
        )
    )

    conn.commit()
    conn.close()

    sync_excel()

    return True, {
        "id": user_id,
        "email": user_email,
    }


# ==========================================================
# DELETE USER
# ==========================================================

def delete_user(user_id):

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute(
        """
        SELECT id, email, created_at
        FROM users
        WHERE id = ?
        """,
        (user_id,)
    )

    user = cursor.fetchone()

    if not user:

        conn.close()

        return False

    # Save deleted user to Excel
    if os.path.exists(EXCEL_PATH):

        wb = load_workbook(EXCEL_PATH)

    else:

        wb = Workbook()

        default_sheet = wb.active
        wb.remove(default_sheet)

    if "Deleted Users" in wb.sheetnames:

        ws = wb["Deleted Users"]

    else:

        ws = wb.create_sheet("Deleted Users")

        ws.append([
            "User ID",
            "Email",
            "Created At",
            "Deleted At",
        ])

    from datetime import datetime

    ws.append([
        user[0],
        user[1],
        user[2],
        datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        ),
    ])

    wb.save(EXCEL_PATH)

    # Delete login history
    cursor.execute(
        """
        DELETE FROM login_activity
        WHERE user_id = ?
        """,
        (user_id,)
    )

    # Delete user
    cursor.execute(
        """
        DELETE FROM users
        WHERE id = ?
        """,
        (user_id,)
    )

    conn.commit()
    conn.close()

    sync_excel()

    return True


# ==========================================================
# CLEAR ALL LOGIN ACTIVITY
# ==========================================================

def clear_login_activity():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM login_activity"
    )

    conn.commit()
    conn.close()

    sync_excel()


# ==========================================================
# STATISTICS
# ==========================================================

def get_total_users():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT COUNT(*) FROM users"
    )

    total = cursor.fetchone()[0]

    conn.close()

    return total


def get_total_logins():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT COUNT(*) FROM login_activity"
    )

    total = cursor.fetchone()[0]

    conn.close()

    return total


# ==========================================================
# USERS
# ==========================================================

def get_users():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            email,
            created_at
        FROM users
        ORDER BY created_at DESC
    """)

    users = cursor.fetchall()

    conn.close()

    return users


# ==========================================================
# LOGIN ACTIVITY
# ==========================================================

def get_login_activity():

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            user_id,
            email,
            login_time
        FROM login_activity
        ORDER BY login_time DESC
    """)

    activity = cursor.fetchall()

    conn.close()

    return activity


# ==========================================================
# OTP
# ==========================================================

OTP_EXPIRY_SECONDS = 300


def generate_otp():

    return str(
        random.randint(100000, 999999)
    )


def send_otp(email, otp):

    api_key = os.getenv("RESEND_API_KEY")

    from_email = os.getenv(
        "RESEND_FROM_EMAIL",
        "onboarding@resend.dev"
    )

    if not api_key:

        return False, "RESEND_API_KEY is missing."

    try:

        resend.api_key = api_key

        resend.Emails.send({
            "from": from_email,
            "to": [email],
            "subject": "CODEX Password Reset OTP",
            "html": f"""
                <div style="font-family: Arial, sans-serif;">
                    <h2>CODEX Password Reset</h2>

                    <p>Your verification code is:</p>

                    <h1>{otp}</h1>

                    <p>
                        This OTP is valid for 5 minutes.
                    </p>

                    <p>
                        If you didn't request a password reset,
                        you can safely ignore this email.
                    </p>
                </div>
            """
        })

        return True, "OTP sent successfully."

    except Exception as e:

        return False, str(e)


# ==========================================================
# RESET PASSWORD
# ==========================================================

def reset_password(email, new_password):

    email = email.strip().lower()

    if len(new_password) < 6:

        return False, (
            "Password must contain at least 6 characters."
        )

    password_hash = hash_password(
        new_password
    )

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute(
        """
        UPDATE users
        SET password_hash = ?
        WHERE email = ?
        """,
        (
            password_hash,
            email,
        ),
    )

    if cursor.rowcount == 0:

        conn.close()

        return False, "User not found."

    conn.commit()
    conn.close()

    sync_excel()

    return True, "Password reset successfully."


# ==========================================================
# INITIALIZE
# ==========================================================

init_db()